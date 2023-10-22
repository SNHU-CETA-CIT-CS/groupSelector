#!/usr/bin/env python
import os
from random import randint
from sys import argv, exit
import platform
from os import path
from logging import basicConfig, getLogger, DEBUG, INFO, CRITICAL
from pickle import dump, load
from time import sleep
from zipfile import ZipFile
from PyQt5 import uic, QtTest
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot, QSettings, Qt, QDir, QCoreApplication
from PyQt5.QtWidgets import QMainWindow, QApplication, QDialog, QMessageBox, QFileSystemModel, QHeaderView
from openpyxl import load_workbook, utils

from uniqueNumGen import UniqueNumGen

# Change the current dir to the temporary one created by PyInstaller for MSWindowOS
try:
    from sys import _MEIPASS
    os.chdir(_MEIPASS)
    print(f"On MSWindows OS {_MEIPASS}")
except ImportError:
    pass    # Must be macOS, just forget it and move on...

import groupSelectorResources_rc

homeDir = path.expanduser('~')

groupSizeDefault = 3
logFilenameDefault = 'groupSelector.log'
rootFolderNameDefault = './'
createLogFileDefault = True
pickleFilenameDefault = ".groupSelectorSavedObjects.pl"
showHelpOnStartupDefault = True

baseDir = os.path.dirname(__file__)


class groupSelector(QMainWindow):
    """A zip file extraction & renaming utility for zyBooks project lab file downloads"""

    def __init__(self, parent=None):
        """ Build a GUI  main window for zipRenamer."""

        super().__init__(parent)
        self.logger = getLogger("Fireheart.groupSelector")
        self.appSettings = QSettings()
        self.quitCounter = 0       # used in a workaround for a QT5 bug.

        self.pickleFilename = pickleFilenameDefault
        self.restoreSettings()

        try:
            self.spreadsheetFileFound, self.statusMessage, self.textOutput = self.restoreApp()
        except FileNotFoundError:
            self.restartApp()

        uic.loadUi("groupSelectorMainWindow.ui", self)

        self.textOutputUI.appendPlainText("Hello")
        self.textOutputUI.repaint()
        self.spreadsheetFileFound = False
        self.statusMessage = ""
        self.columnNames = ["" for count in range(0, 10)]
        self.studentNames = []
        self.uiCheckboxes = [self.checkBox1UI,
                             self.checkBox2UI,
                             self.checkBox3UI,
                             self.checkBox4UI,
                             self.checkBox5UI,
                             self.checkBox6UI,
                             self.checkBox7UI,
                             self.checkBox8UI,
                             self.checkBox9UI,
                             self.checkBox10UI,
                            ]
        self.columnChecked = [False for column in range(0, 10)]
        self.hasHeaderRow = True
        # self.groupSize = 0
        if self.logger.getEffectiveLevel() == 10:   # Logger is set to debug level.
            self.textOutput = f"createLogFile: {self.createLogFile}\nspreadsheetFilename: {self.spreadsheetFilename}\nlogFilename: {self.logFilename}\ngroupSize: {self.groupSize}\npickleFilename: {self.pickleFilename}\nshowHelpOnStartup: {self.showHelpOnStartup}"
        else:
            self.textOutput = ""

        self.preferencesSelectButton.clicked.connect(self.preferencesSelectButtonClickedHandler)
        self.helpSelectUI.clicked.connect(self.helpSelectButtonClickedHandler)
        self.spreadsheetSelectButton.clicked.connect(self.spreadsheetSelectButtonClickedHandler)
        self.currentSpreadsheetPathLabel.clicked.connect(self.spreadsheetSelectButtonClickedHandler)
        self.generateGroupsButton.clicked.connect(self.createGroupsButtonClickedHandler)
        self.hasHeaderRowUI.toggled.connect(self.headerButtonHandler)
        self.noHeaderRowUI.toggled.connect(self.noHeaderButtonHandler)

        self.setWindowIcon(QIcon('images/groupSelector.png'))
        self.logger.info("Application startup completed.")

        self.updateUI()
        # Startup with the help dialog opened.
        if self.showHelpOnStartup:
            self.helpSelectButtonClickedHandler()

    def __str__(self):
        """String representation for groupSelector.
        """
        return "A student random group creation utility"

    def updateUI(self):
        self.textOutputUI.setPlainText(self.textOutput)
        self.textOutputUI.repaint()
        self.currentSpreadsheetPathLabel.setText(self.spreadsheetFilename)
        if len(self.statusMessage) > 0:
            self.gSelectorStatusBarUI.showMessage(self.statusMessage, 5000)
            self.statusMessage = ""
        for boxNumber, checkBox in enumerate(self.uiCheckboxes):
            checkBox.setText(self.columnNames[boxNumber])
            if "name" in self.columnNames[boxNumber].lower():
                checkBox.setChecked(True)
                self.columnChecked[boxNumber] = True
            else:
                checkBox.setChecked(False)
                self.columnChecked[boxNumber] = False
        if self.hasHeaderRow:
            self.hasHeaderRowUI.setChecked(True)
        else:
            self.hasHeaderRowUI.setChecked(False)

    def setSpreadsheetFilename(self, newFilename):
        self.spreadsheetFilename = newFilename
        self.appSettings.setValue('spreadsheetFilename', self.spreadsheetFilename)

    def getSpreadsheetFilename(self):
        return self.spreadsheetFilename

    def restartApp(self):
        if self.createLogFile:
            self.logger.debug("Restarting program")

    def saveApp(self):
        if self.createLogFile:
            self.logger.debug("Saving program state")
        saveItems = (self.spreadsheetFileFound, self.statusMessage, self.textOutput)
        if self.appSettings.contains('pickleFilename'):
            with open(path.join(path.dirname(path.realpath(__file__)),  self.appSettings.value('pickleFilename', type=str)), 'wb') as pickleFile:
                dump(saveItems, pickleFile)
        elif self.createLogFile:
            self.logger.critical("No pickle Filename")

    def restoreApp(self):
        if self.appSettings.contains('pickleFilename'):
            self.appSettings.value('pickleFilename', type=str)
            with open(path.join(path.dirname(path.realpath(__file__)),  self.appSettings.value('pickleFilename', type=str)), 'rb') as pickleFile:
                return load(pickleFile)
        else:
            self.logger.critical("No pickle Filename")

    def restoreSettings(self):
        # Restore settings values, write defaults to any that don't already exist.
        if appSettings.contains('createLogFile'):
            self.createLogFile = appSettings.value('createLogFile')
        else:
            self.createLogFile = createLogFileDefault
            appSettings.setValue('createLogFile', self.createLogFile)

        if self.createLogFile:
            self.logger.debug("Starting restoreSettings")

        if self.appSettings.contains('spreadsheetFilename'):
            self.spreadsheetFilename = self.appSettings.value('spreadsheetFilename', type=str)
        else:
            self.spreadsheetFilename = rootFolderNameDefault
            self.appSettings.setValue('spreadsheetFilename', self.spreadsheetFilename)

        if self.appSettings.contains('logFilename'):
            self.logFilename = self.appSettings.value('logFilename', type=str)
        else:
            self.logFilename = logFilenameDefault
            self.appSettings.setValue('logFilename', self.logFilename)

        if self.appSettings.contains('groupSize'):
            self.groupSize = self.appSettings.value('groupSize', type=int)
        else:
            self.groupSize = groupSizeDefault
            self.appSettings.setValue('groupSize', self.groupSize)

        if self.appSettings.contains('pickleFilename'):
            self.pickleFilename = self.appSettings.value('pickleFilename', type=str)
        else:
            self.pickleFilename = pickleFilenameDefault
            self.appSettings.setValue('pickleFilename', self.pickleFilename)

        if self.appSettings.contains('showHelpOnStartup'):
            self.showHelpOnStartup = self.appSettings.value('showHelpOnStartup', type=bool)
        else:
            self.showHelpOnStartup = showHelpOnStartupDefault
            self.appSettings.setValue('showHelpOnStartup', self.showHelpOnStartup)

    # self.groupSizeSpinUI.valueChanged.connect(self.nameOnlySelected)

    def getStudentNames(self, spreadsheetFilename):
        studentList = []
        try:
            workbook = load_workbook(filename=spreadsheetFilename)
            sheet = workbook.active
            if self.hasHeaderRow:
                startingRow = 2
            else:
                startingRow = 1
            for rowNumber in range(startingRow, sheet.max_row + 1):
                studentName = []
                for columnNumber in range(1, sheet.max_column + 1):
                    if self.columnChecked[columnNumber - 1]:
                        cellObject = sheet.cell(row=rowNumber, column=columnNumber)
                        studentName.append(cellObject.value)
                        # print(f"{cellObject.value}", end='')
                studentList.append(" ".join(studentName))
                # print()
        except FileNotFoundError:
            print(f"fCould not open {spreadsheetFilename}")
        return studentList

    def updateColumnSelections(self, spreadsheetFilename):
        try:
            workbook = load_workbook(filename=spreadsheetFilename)
            sheet = workbook.active
            self.columnNames = []
            for cellNumber, cell in enumerate(sheet[1]):
                if cell.value is None or cellNumber >= 11:
                    break
                self.columnNames.append(cell.value)
            for columnNumber in range(cellNumber + 1, 11):  # Fill in the 'empty' cell names with empty strings.
                self.columnNames.append("")
        except FileNotFoundError:
            print(f"fCould not open {spreadsheetFilename}")
        self.updateUI()

    def generateStudentGroups(self, listOfStudents):
        studentPicker = UniqueNumGen(0, len(listOfStudents) - 1)
        lastGroupNumber = 0
        numberOfStudents = len(listOfStudents)
        numberOfGroups = numberOfStudents // self.groupSize
        if numberOfStudents % self.groupSize != 0:
            numberOfGroups += 1

        self.textOutput += f"\nWith {numberOfStudents} students in the class, there will be {numberOfGroups} groups"
        # print(f"\nWith {numberOfStudents} students in the class, there will be {numberOfGroups} groups")
        for groupNumber in range(1, self.groupSize + 1):
            self.textOutput += f"\nGroup {groupNumber} members are...\n"
            # print(f"\nGroup {groupNumber} members are...")
            lastGroupNumber = groupNumber + 1
            for studentNumber in range(0, self.groupSize):
                nextStudent = listOfStudents[studentPicker.getNext()]
                self.textOutput += f"\t{nextStudent}\n"
                # print(f"\t{nextStudent}")
                delay = randint(3, 6) * 1000
                QtTest.QTest.qWait(delay)
                self.updateUI()

        self.textOutput += f"\nGroup {lastGroupNumber} members are...\n"
        # print(f"\nGroup {lastGroupNumber} members are...")
        for remainingStudents in studentPicker.getUnused():
            nextStudent = listOfStudents[studentPicker.getNext()]
            self.textOutput += f"\t{nextStudent}\n"
            # print(nextStudent)
            delay = randint(3, 6) * 1000
            QtTest.QTest.qWait(delay)
            self.updateUI()

    def createGroupsButtonClickedHandler(self):
        filename = self.getSpreadsheetFilename()
        if not path.exists(filename):
            self.statusMessage = f"File {filename} doesn't exist. Click the Spreadsheet Icon to select a different one."
        self.textOutput = f"Reading student names from spreadsheet named:\n  {filename}:\n"
        self.studentNames = self.getStudentNames(filename)
        self.generateStudentGroups(self.studentNames)
        self.updateUI()

    @pyqtSlot()  # User is requesting a top level folder select.
    def headerButtonHandler(self):
        self.hasHeaderRow = True

    @pyqtSlot()  # User is requesting a top level folder select.
    def noHeaderButtonHandler(self):
        self.hasHeaderRow = False

    @pyqtSlot()  # User is requesting a top level folder select.
    def spreadsheetSelectButtonClickedHandler(self):
        filename = self.getSpreadsheetFilename()
        folderDialog = FileSelectDialog(filename)
        folderDialog.show()
        folderDialog.exec_()
        self.updateColumnSelections(filename)
        self.updateUI()

    @pyqtSlot()  # User is requesting preferences editing dialog box.
    def preferencesSelectButtonClickedHandler(self):
        if self.createLogFile:
            self.logger.info("Setting preferences")
        preferencesDialog = PreferencesDialog()
        preferencesDialog.show()
        preferencesDialog.exec_()
        self.restoreSettings()              # 'Restore' settings that were changed in the dialog window.
        self.updateUI()

    @pyqtSlot()  # User is requesting preferences editing dialog box.
    def helpSelectButtonClickedHandler(self):
        if self.createLogFile:
            self.logger.info("Help dialog opened.")
        helpDialog = HelpDialog()
        helpDialog.show()
        helpDialog.exec_()
        self.restoreSettings()              # 'Restore' settings that were changed in the dialog window.
        self.updateUI()

    @pyqtSlot()				# Player asked to quit the game.
    def closeEvent(self, event):
        if self.createLogFile:
            self.logger.debug("Closing app event")
        if self.quitCounter == 0:
            self.quitCounter += 1
            quitMessage = "Are you sure you want to quit?"
            reply = QMessageBox.question(self, 'Message', quitMessage, QMessageBox.Yes, QMessageBox.No)

            if reply == QMessageBox.Yes:
                self.saveApp()
                event.accept()
            else:
                self.quitCounter = 0
                event.ignore()


class FileSelectDialog(QDialog):
    def __init__(self, startingFolderName, parent=groupSelector):
        super(FileSelectDialog, self).__init__()
        uic.loadUi('rootSelectDialog.ui', self)
        if platform.system() == "Darwin+":
            pass
        else:
            fileModel = QFileSystemModel()
            fileModel.setFilter(QDir.AllDirs | QDir.AllEntries | QDir.NoDotAndDotDot)
            fileModel.setNameFilters(['*.xlsx'])
            fileModel.setRootPath(startingFolderName)
            self.selectedFile = '/Volumes/Macintosh HD'
            self.selectedFile = startingFolderName
            self.fileSelectTreeView.setModel(fileModel)
            self.fileSelectTreeView.setCurrentIndex(fileModel.index(startingFolderName))
            self.fileSelectTreeView.header().setSectionResizeMode(0, QHeaderView.ResizeToContents)

        if platform.system() == "Darwin+":
            iconFilename = 'images/groupSelector.icns'
        else:
            iconFilename = 'images/groupSelector.ico'
        self.setWindowIcon(QIcon(os.path.join(baseDir, iconFilename)))

        self.fileSelectTreeView.doubleClicked.connect(self.fileDoubleClickedHandler)
        self.buttonBox.rejected.connect(self.cancelClickedHandler)
        self.buttonBox.accepted.connect(self.okayClickedHandler)
        self.fileSelectTreeView.clicked.connect(self.selectionChangedHandler)
        self.fileSelectTreeView.expanded.connect(self.itemExpandedHandler)

    # @pyqtSlot()
    def fileDoubleClickedHandler(self, signal):
        # print(signal)
        filename = self.fileSelectTreeView.model().filePath(signal)
        if path.isfile(filename):
            # print(filePath)
            GroupSelectorApp.setSpreadsheetFilename(filename)
            self.close()
        else:
            print("File selected, not an .xlsx file")

    # @pyqtSlot()
    def okayClickedHandler(self):
        # print(self.selectedFile)
        if path.isfile(self.selectedFile):
            GroupSelectorApp.setSpreadsheetFilename(self.selectedFile)
            self.close()
        else:
            print("File Selected on Okay")

    # @pyqtSlot(QItemSelection)
    def selectionChangedHandler(self, selected):
        self.fileSelectTreeView.resizeColumnToContents(0)
        # print(selected)
        if path.isdir(self.fileSelectTreeView.model().filePath(selected)):
            self.selectedFile = self.fileSelectTreeView.model().filePath(selected)
            # print(self.selectedFile)
        else:
            print("File Selected")

    # @pyqtSlot()
    def itemExpandedHandler(self, selected):
        self.fileSelectTreeView.resizeColumnToContents(0)

    # @pyqtSlot()
    def cancelClickedHandler(self):
        self.close()


class PreferencesDialog(QDialog):
    def __init__(self, parent=groupSelector):
        super(PreferencesDialog, self).__init__()

        uic.loadUi('preferencesDialog.ui', self)
        self.logger = getLogger("Fireheart.groupSelector")

        self.logger.debug("Starting Preferences Dialog launch.")
        self.appSettings = QSettings()
        if self.appSettings.contains('logFilename'):
            self.logFilename = self.appSettings.value('logFilename', type=str)
        else:
            self.logFilename = logFilenameDefault
            self.appSettings.setValue('logFilename', self.logFilename)

        if self.appSettings.contains('groupSize'):
            self.groupSize = self.appSettings.value('groupSize', type=int)
        else:
            self.groupSize = groupSizeDefault
            self.appSettings.setValue('groupSize', self.groupSize)

        if self.appSettings.contains('createLogFile'):
            self.createLogFile = self.appSettings.value('createLogFile')
        else:
            self.createLogFile = logFilenameDefault
            self.appSettings.setValue('createLogFile', self.createLogFile)
        self.logger.debug("Preferences settings restored.")

        if platform.system() == "Darwin+":
            iconFilename = 'images/groupSelector.icns'
        else:
            iconFilename = 'images/groupSelector.ico'
        self.setWindowIcon(QIcon(os.path.join(baseDir, iconFilename)))

        self.buttonBox.rejected.connect(self.cancelClickedHandler)
        self.buttonBox.accepted.connect(self.okayClickedHandler)
        self.logFilenameUI.editingFinished.connect(self.logFilenameChanged)
        self.groupSizeDefaultUI.editingFinished.connect(self.logFilenameChanged)
        self.createLogFileUI.stateChanged.connect(self.createLogFileChanged)
        self.logger.debug("Preferences dialog object built.")

        self.updateUI()

    # @pyqtSlot()
    def logFilenameChanged(self):
        self.logFilename = self.logFilenameUI.text()

    # @pyqtSlot()
    def groupSizeChanged(self):
        self.groupSize = int(self.groupSize.text())

    # @pyqtSlot()
    def createLogFileChanged(self):
        self.createLogFile = self.createLogFileCUI.isChecked()

    def updateUI(self):
        self.logger.debug("Updating Preferences UI.")
        self.logFilenameUI.setText(str(self.logFilename))
        self.groupSizeDefaultUI.setText(str(self.groupSize))

        if self.createLogFile:
            self.createLogFileUI.setCheckState(Qt.Checked)
        else:
            self.createLogFileUI.setCheckState(Qt.Unchecked)
        self.logger.debug("Preferences UI Updated.")

    # @pyqtSlot()
    def okayClickedHandler(self):
        # write out all settings
        preferencesGroup = (('logFilename', self.logFilename),
                            ('groupSize', self.groupSize),
                            ('createLogFile', self.createLogFile),
                            )
        # Write settings values.
        for setting, variableName in preferencesGroup:
            self.appSettings.setValue(setting, variableName)
        self.close()

    # @pyqtSlot()
    def cancelClickedHandler(self):
        self.close()


class HelpDialog(QDialog):
    def __init__(self, parent=groupSelector):
        super(HelpDialog, self).__init__()

        uic.loadUi('helpDialog.ui', self)
        self.logger = getLogger("Fireheart.groupSelector")

        self.appSettings = QSettings()
        if self.appSettings.contains('showHelpOnStartup'):
            self.showHelpOnStartup = self.appSettings.value('showHelpOnStartup', type=bool)
        else:
            self.self.showHelpOnStartup = showHelpOnStartupDefault
            self.appSettings.setValue('showHelpOnStartup', self.showHelpOnStartup)

        if platform.system() == "Darwin+":
            iconFilename = 'images/groupSelector.icns'
        else:
            iconFilename = 'images/groupSelector.ico'
        self.setWindowIcon(QIcon(os.path.join(baseDir, iconFilename)))

        self.buttonBox.accepted.connect(self.okayClickedHandler)
        self.helpOnStartupUI.stateChanged.connect(self.helpOnStartupChanged)

        try:
            self.helpText = ""
            with open("helpDialog.html", 'r') as helpTextFile:
                for line in helpTextFile.readlines():
                    self.helpText += line.strip()
        except FileNotFoundError:
            self.helpText = "Help Text file is missing!"

        self.updateUI()

    # @pyqtSlot()
    def updateUI(self):
        self.helpTextUI.setText(self.helpText)

        if not self.showHelpOnStartup:
            self.helpOnStartupUI.setCheckState(Qt.Checked)
        else:
            self.helpOnStartupUI.setCheckState(Qt.Unchecked)

    # @pyqtSlot()
    def helpOnStartupChanged(self):
        self.showHelpOnStartup = not self.helpOnStartupUI.isChecked()

    # @pyqtSlot()
    def okayClickedHandler(self):
        # write out all settings
        helpGroup = (('showHelpOnStartup', self.showHelpOnStartup),
                     )
        # Write settings values.
        for setting, variableName in helpGroup:
            self.appSettings.setValue(setting, variableName)
        self.close()

    # @pyqtSlot()
    def cancelClickedHandler(self):
        self.close()


if __name__ == "__main__":
    QCoreApplication.setOrganizationName("Fireheart Software")
    QCoreApplication.setOrganizationDomain("fireheartsoftware.com")
    QCoreApplication.setApplicationName("groupSelector")
    appSettings = QSettings()
    if appSettings.contains('createLogFile'):
        createLogFile = appSettings.value('createLogFile')
    else:
        createLogFile = createLogFileDefault
        appSettings.setValue('createLogFile', createLogFile)

    if createLogFile:
        startingFolderName = path.dirname(path.realpath(__file__))
        if appSettings.contains('logFilename'):
            logFilename = appSettings.value('logFilename', type=str)
        else:
            logFilename = logFilenameDefault
            appSettings.setValue('logFilename', logFilename)
        basicConfig(filename=path.join(startingFolderName, logFilename), level=INFO,
                    format='%(asctime)s %(name)-8s %(levelname)-8s %(message)s')
    app = QApplication(argv)
    if platform.system() == "Darwin+":
        iconFilename = 'images/groupSelector.icns'
    else:
        iconFilename = 'images/groupSelector.ico'

    app.setWindowIcon(QIcon(os.path.join(baseDir, iconFilename)))
    GroupSelectorApp = groupSelector()
    GroupSelectorApp.updateUI()
    GroupSelectorApp.show()
    exit(app.exec_())
